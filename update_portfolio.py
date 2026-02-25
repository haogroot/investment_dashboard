import pandas as pd
import yfinance as yf
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import sys
import os
from pathlib import Path
from datetime import datetime, timedelta
import argparse

# --- Configuration ---
BASE_DIR = Path(__file__).parent
INPUT_FILE = BASE_DIR / "trade_source" / "fubon-trade-record_20260224.csv"
OUTPUT_DIR = BASE_DIR / "trade_output"
# Find the latest output file (assuming pattern portfolio_dashboard_YYYYMMDD.xlsx)
# We will search for the latest file to update.

def get_latest_dashboard_file():
    files = list(OUTPUT_DIR.glob("portfolio_dashboard_*.xlsx"))
    if not files:
        return None
    # Sort by name (date) descending
    files.sort(key=lambda x: x.name, reverse=True)
    return files[0]

def update_portfolio(mode='full', manual_trade=None):
    target_file = get_latest_dashboard_file()
    if not target_file:
        print("No existing portfolio dashboard found. Please run 'build_portfolio.py' first.")
        return

    print(f"Updating: {target_file.name}")
    print(f"Mode: {mode}")

    # 1. Open Workbook
    try:
        wb = openpyxl.load_workbook(target_file)
    except Exception as e:
        print(f"Error opening workbook: {e}")
        return

    # 2. Determine Last Date and State
    ws_market = wb['Market_Data']
    # Assuming Date is in Column A (Index 1) and headers in Row 1
    # Find last row
    last_row = ws_market.max_row
    last_date_val = ws_market.cell(row=last_row, column=1).value
    
    if isinstance(last_date_val, str):
        # clean timestamp if string "YYYY-MM-DD ..."
        last_date_val = pd.to_datetime(last_date_val).normalize()
    elif isinstance(last_date_val, datetime):
        last_date_val = pd.to_datetime(last_date_val).normalize()
    else:
        print(f"Could not determine last date. Found: {last_date_val}")
        return

    print(f"Last update date: {last_date_val.date()}")
    
    today = pd.Timestamp.now().normalize()
    # Logic moved to Step 5 (Fetch Market Data) to handle Overwrite vs Skip
    # if last_date_val >= today: ... (Removed)

    # 3. Identify Tickers
    # Read tickers from header of Market_Data (Row 1), skipping Date (Col 1) and SP500 (Last Col probably)
    # Actually, let's just use Reference_Data to get purely portfolio tickers
    ws_ref = wb['Reference_Data']
    ref_rows = list(ws_ref.values)
    ref_header = ref_rows[0]
    ref_data = ref_rows[1:]
    # Ticker is usually col 0
    portfolio_tickers = [row[0] for row in ref_data]
    
    # Check headers in Market_Data to find map
    md_header = [c.value for c in ws_market[1]]
    # md_header[0] is Date. 
    # Valid tickers in market data (includes SP500)
    market_tickers = md_header[1:]
    
    # 4. Handle Trades
    # If manual trade, append to CSV first
    if mode == 'trade' and manual_trade:
        print(f"Appending manual trade: {manual_trade}")
        # Format: Date, Ticker, Side, Price, Qty, Fee, Tax
        # manual_trade arg is list: [Date, Ticker, Side, Qty, Price]
        # We need to match CSV format
        # Check CSV encoding
        try:
            with open(INPUT_FILE, 'r', encoding='utf-8') as f:
                header = f.readline().strip().split(',')
        except:
             with open(INPUT_FILE, 'r', encoding='big5') as f:
                header = f.readline().strip().split(',')
        
        # We assume standard 7 params or use default for fee/tax
        # User input: '2026-01-27 AAPL BUY 100 150.50'
        # CSV Order: Date, Ticker, Side, Price, Qty, Fee, Tax
        t_date, t_ticker, t_side, t_qty, t_price = manual_trade
        
        # Check if trade date is in the past relative to dashboard
        trade_dt = pd.to_datetime(t_date).normalize()
        if trade_dt <= last_date_val:
            print(f"WARNING: You are adding a trade for {t_date}, but the dashboard is already updated to {last_date_val.date()}.")
            print("This update script only APPENDS new days. It cannot modify historical unit counts.")
            print("Please run 'python build_portfolio.py' to rebuild the portfolio with correct history.")
            return

        new_row = f"{t_date},{t_ticker},{t_side},{t_price},{t_qty},0,0\n"
        
        with open(INPUT_FILE, 'a', encoding='utf-8') as f:
            f.write(new_row)
        print("Trade appended to CSV.")

    # 5. Fetch New Market Data
    # Logic for Overwrite vs Append
    # We always look back 7 days to ensure recent missing data (e.g. yesterday) is filled,
    # and to handle the "today" overwrite case naturally.
    lookback_days = 3
    start_fetch_date = last_date_val - timedelta(days=lookback_days)
    
    # Check manual trade backdate - logic complicates if we look back.
    # For simplicity, manual trades still append to CSV.
    # The portfolio update logic below handles Market Data / Units for *fetched* days.
    # If a manual trade is added for a day within lookback, it MIGHT be processed if we re-calc units from CSV?
    # Current script uses `current_units` from Excel state.
    # If we want to support backdated trades properly, we'd need to re-run from CSV.
    # But for now, let's focus on fixing Market Data gaps.
    
    fetch_end_date = today + timedelta(days=1)
    
    fetch_days = pd.date_range(start_fetch_date, today, freq='B')
    
    # Determine if today's data is being overwritten
    overwrite_today = (last_date_val == today)

    if len(fetch_days) == 0:
         print("No relevant days to update.")
    else:
        print(f"Fetching data from {start_fetch_date.date()} to {today.date()} (Lookback {lookback_days} days)")
        
        # Fetch data
        to_fetch = [t if t != 'SP500' else '^GSPC' for t in market_tickers]
        
        try:
            # Fetch with explicit start/end
            new_mkt_data = yf.download(to_fetch, start=start_fetch_date, end=fetch_end_date, threads=True)['Close']
            if isinstance(new_mkt_data, pd.Series):
                new_mkt_data = new_mkt_data.to_frame()
            
            # Ensure index is timezone-naive
            if new_mkt_data.index.tz is not None:
                new_mkt_data.index = new_mkt_data.index.tz_localize(None)

            # Map ticker -> col index in Excel
            ticker_col_map = {t: i+2 for i, t in enumerate(market_tickers)}
            
            # Helper to find row for date
            # We scan the last 30 rows for efficiency? Or just scanning all is fine for small sheets.
            # Market_Data sorted by date? Yes.
            # We can map Date -> Row Index from existing sheet
            existing_dates = {}
            for row in ws_market.iter_rows(min_row=2, max_col=1, values_only=False):
                cell = row[0]
                val = cell.value
                if val:
                    if isinstance(val, str):
                        try: val = pd.to_datetime(val).normalize()
                        except: continue
                    elif isinstance(val, datetime):
                        val = pd.to_datetime(val).normalize()
                    existing_dates[val] = cell.row
            
            for date in fetch_days:
                # 1. Determine Row Index (Overwrite or Append)
                if date in existing_dates:
                    target_row = existing_dates[date]
                    action = "Overwrite"
                else:
                    target_row = ws_market.max_row + 1
                    action = "Append"
                    existing_dates[date] = target_row # Update map
                
                # Write Date
                ws_market.cell(row=target_row, column=1, value=date)
                
                # Write Prices
                if date in new_mkt_data.index:
                    row_data = new_mkt_data.loc[date]
                    found_data = False
                    for tick, price in row_data.items():
                        if tick in ticker_col_map:
                            col_idx = ticker_col_map[tick]
                            if pd.notnull(price):
                                ws_market.cell(row=target_row, column=col_idx, value=price)
                                found_data = True
                    # if found_data: print(f"  Updated {date.date()} ({action})")
                
                # If Append, we need to add Formulas to other sheets
                if action == "Append":
                    # --- Daily_Returns ---
                    ws_ret = wb['Daily_Returns']
                    ws_ret_row = ws_ret.max_row + 1
                    ws_ret.cell(row=ws_ret_row, column=1, value=f'=Market_Data!A{target_row}').number_format = 'yyyy-mm-dd'
                    for tick in market_tickers:
                        if tick in ticker_col_map:
                            c = ticker_col_map[tick]
                            col_letter = get_column_letter(c)
                            formula = f'=LN(Market_Data!{col_letter}{target_row}/Market_Data!{col_letter}{target_row-1})'
                            ws_ret.cell(row=ws_ret_row, column=c, value=formula).number_format = '0.00%'

                    # --- Daily_Drawdowns ---
                    ws_dd = wb['Daily_Drawdowns']
                    ws_dd_row = ws_dd.max_row + 1
                    ws_dd.cell(row=ws_dd_row, column=1, value=f'=Market_Data!A{target_row}').number_format = 'yyyy-mm-dd'
                    for tick in market_tickers:
                        if tick in ticker_col_map:
                            c = ticker_col_map[tick]
                            col_letter = get_column_letter(c)
                            formula = f'=Market_Data!{col_letter}{target_row}/MAX(Market_Data!{col_letter}$2:Market_Data!{col_letter}{target_row}) - 1'
                            ws_dd.cell(row=ws_dd_row, column=c, value=formula).number_format = '0.00%'

        except Exception as e:
            print(f"Error fetching market data: {e}")
            return


    # 7. Update Daily_Units and Daily_Cash
    if mode == 'quick':
        print("Skipping trade logic (Quick Mode).")
        # Ensure if overwriting, we don't need to do anything for units/cash unless we want to support 'quick' trade corrections?
        # But for quick mode, we assume units/cash are static.
        # So if overwrite, we essentially "restore" valid state or just leave it?
        # If we re-write row with "yesterday's units", we effectively undo today's trades if we don't process them.
        # SAFE PATH: In Quick Mode + Overwrite, DO NOT TOUCH Units/Cash.
    else:
        # Check for new trades!
        # Load CSV
        try:
           df_trades_all = pd.read_csv(INPUT_FILE)
           df_trades_all['Date'] = pd.to_datetime(df_trades_all['Date']).dt.normalize()
        except:
           print("Error reading trade file.")
           return

        # Check for trades AFTER last_date_val (exclusive) OR equal if overwrite?
        # If overwrite_today, we need to check trades for TODAY.
        if overwrite_today:
             # Find trades == today
             # But wait, original build_portfolio already processed today's trades?
             # Yes, if build_portfolio ran today, it processed today's trades.
             # So unless user ADDED a trade to CSV *after* build, units are already correct.
             # If we want to support "Add Trade today and update", we need to RE-CALCULATE today's units?
             # That's complex because we need "Yesterday's units".
             # Daily_Units last row = Today's Units.
             # If we overwrite, we need Units[Yesterday] + Trades[Today].
             # We can get Units[Yesterday] from (last_row - 1).
             new_trades = df_trades_all[df_trades_all['Date'] >= last_date_val]
        else:
             new_trades = df_trades_all[df_trades_all['Date'] > last_date_val]
        
    ws_units = wb['Daily_Units']
    ws_cash = wb['Daily_Cash']
    
    # Map ticker to Unit Sheet Column
    unit_header = [c.value for c in ws_units[1]]
    unit_col_map = {t: i+1 for i, t in enumerate(unit_header) if i > 0} # Col 1 is Date

    # Get last known state
    last_unit_row = ws_units.max_row
    last_cash_row = ws_cash.max_row # Should be same
    
    # To handle overwrite correctly, if overwrite_today, we need state from Yesterday (row-1)
    if overwrite_today and last_unit_row > 2: # Ensure there is history
         prev_unit_row = last_unit_row - 1
         current_units = {}
         for t, c in unit_col_map.items():
            current_units[t] = ws_units.cell(row=prev_unit_row, column=c).value or 0
         
         current_cash = ws_cash.cell(row=prev_unit_row, column=2).value
    else:
         current_units = {}
         for t, c in unit_col_map.items():
            current_units[t] = ws_units.cell(row=last_unit_row, column=c).value or 0
         current_cash = ws_cash.cell(row=last_cash_row, column=2).value # Cash is Col 2 (B)
    
    # Iterate through new days (fetch_days)
    if len(fetch_days) > 0:
        for date in fetch_days:
            is_today_overwrite = (overwrite_today and date == today_dt)

            if mode == 'quick': 
                 # Just carry forward previous day's units/cash
                 # If overwrite, we just keep yesterday's units? 
                 # No, if build_portfolio ran today, it established today's units.
                 # If quick update, we assume units unchanged.
                 # So if overwrite, we essentially "restore" valid state or just leave it?
                 # If we re-write row with "yesterday's units", we effectively undo today's trades if we don't process them.
                 # SAFE PATH: In Quick Mode + Overwrite, DO NOT TOUCH Units/Cash.
                 pass
            else:
                # 1. Apply trades for this date
                day_trades = new_trades[new_trades['Date'] == date] if not new_trades.empty else pd.DataFrame()
                
                for _, trade in day_trades.iterrows():
                    t_tick = trade['Ticker']
                    t_side = str(trade['Side']).strip().lower()
                    t_qty = float(trade['Qty'])
                    t_price = float(trade['Price'])
                    t_fee = float(trade['Fee']) if pd.notnull(trade['Fee']) else 0
                    t_tax = float(trade['Tax']) if pd.notnull(trade['Tax']) else 0
                    total = t_qty * t_price
                    
                    # Check for New Ticker
                    if t_tick not in unit_col_map:
                        print(f"WARNING: New ticker found: {t_tick}. This script cannot handle new tickers dynamically.")
                        print("Please run 'python build_portfolio.py' for a full rebuild.")
                        return 

                    if 'buy' in t_side or '買' in t_side:
                        current_cash -= (total + t_fee)
                        current_units[t_tick] += t_qty
                    elif 'sell' in t_side or '賣' in t_side:
                        current_cash += (total - t_fee - t_tax)
                        current_units[t_tick] -= t_qty
                
                # 2. Write Row to Daily_Units / Daily_Cash
                if is_today_overwrite:
                    u_row = last_unit_row
                    c_row = last_cash_row
                else:
                    u_row = ws_units.max_row + 1
                    c_row = ws_cash.max_row + 1

                # Units
                ws_units.cell(row=u_row, column=1, value=date) # Date in A
                for t, qty in current_units.items():
                    c = unit_col_map[t]
                    ws_units.cell(row=u_row, column=c, value=qty)
                
                # Cash
                ws_cash.cell(row=c_row, column=1, value=date) # Date in A
                ws_cash.cell(row=c_row, column=2, value=current_cash) # Cash in B

    # 8. Extend Equity_Curve Formulas
    # If overwrite_today, formulas already exist in the last row. No need to add.
    if not overwrite_today:
        ws_eq = wb['Equity_Curve']
        # If we added N days, we need to add N rows to Equity_Curve
        # Start from: last_row + 1. End at: last_row + len(fetch_days)
        
        start_eq_row = ws_eq.max_row + 1
        
        for i, date in enumerate(fetch_days):
            row = start_eq_row + i # Excel Row Index
            prev_row = row - 1
            
            # Col A: Date formula =Market_Data!A{row}
            ws_eq.cell(row=row, column=1, value=f'=Market_Data!A{row}').number_format = 'yyyy-mm-dd'
            
            # Col B: Cash =Daily_Cash!B{row}
            ws_eq.cell(row=row, column=2, value=f'=Daily_Cash!B{row}').number_format = '$#,##0.00'
            
            # Col C: Invested Value. SUMPRODUCT(Daily_Units!B{row}:Z{row}, Market_Data!B{row}:Z{row})
            # Need correct last column letter.
            last_col_idx = len(unit_col_map) + 1 # Date(1) + Tickers...
            last_col_letter = get_column_letter(last_col_idx)
            
            formula_iv = f'=SUMPRODUCT(Daily_Units!B{row}:{last_col_letter}{row},Market_Data!B{row}:{last_col_letter}{row})'
            ws_eq.cell(row=row, column=3, value=formula_iv).number_format = '$#,##0.00'
            
            # Col D: Total NAV =B{row}+C{row}
            ws_eq.cell(row=row, column=4, value=f'=B{row}+C{row}').number_format = '$#,##0.00'
            
            # Col E: Daily Return =(D{row}-D{prev})/D{prev}
            ws_eq.cell(row=row, column=5, value=f'=(D{row}-D{prev_row})/D{prev_row}').number_format = '0.00%'
            
            # Col F: Cum Return =(D{row}-Initial)/Initial
            # Initial is in D2 usually? Or define constant?
            # Let's refer to D2. Assuming D2 is start.
            ws_eq.cell(row=row, column=6, value=f'=(D{row}-D2)/D2').number_format = '0.00%'
            
            # Col G: Drawdown =D{row}/MAX(D$2:D{row})-1
            ws_eq.cell(row=row, column=7, value=f'=D{row}/MAX(D$2:D{row})-1').number_format = '0.00%'


    print("Saving workbook...")
    wb.save(target_file)
    print("Update complete.")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Update Portfolio Dashboard')
    parser.add_argument('mode', nargs='?', default='full', help='Update mode: full, quick, trade')
    parser.add_argument('trade_args', nargs='*', help='Args for trade mode: Date Ticker Side Qty Price')
    
    args = parser.parse_args()
    
    manual_trade = None
    if args.mode == 'trade':
        if len(args.trade_args) < 5:
            print("Usage: python update_portfolio.py trade YYYY-MM-DD TICKER SIDE QTY PRICE")
            sys.exit(1)
        manual_trade = args.trade_args
        
    update_portfolio(args.mode, manual_trade)
